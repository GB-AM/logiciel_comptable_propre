# app.py → Colle ÇA à la place de tout
from flask import Flask, request, send_file, jsonify
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

app = Flask(__name__)

@app.route('/')
def home():
    return "Robot certificat prêt ! Envoie-moi du JSON 🍀"

@app.route('/certificat', methods=['POST'])
def certificat():
    data = request.json  # Base44 t’envoie ça
    entreprise = data['entreprise']
    montant = float(data['montant'])
    date = data['date']
    penalite = float(data.get('penalite', 0))
    cie = float(data.get('cie', 0))

    # Ouvre TON Excel
    wb = load_workbook('Modèle de ce qui est attendu.xlsx')
    ws = wb['03']

    # Trouve la prochaine ligne vide
    row = ws.max_row + 1
    ws[f'A{row}'] = date
    ws[f'C{row}'] = montant
    ws[f'D{row}'] = montant * 0.05  # retenue 5%
    ws[f'E{row}'] = penalite
    ws[f'G{row}'] = cie
    ws[f'H{row}'] = montant * 1.20  # TTC

    # Sauvegarde dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Certificat_{entreprise}_{date}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)
