# fichier backend.py (à coller sur Render.com)
from flask import Flask, request, send_file
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

app = Flask(__name__)

@app.route('/certificat', methods=['POST'])
def genere_certificat():
    data = request.json  # ← Base44 t’envoie ça en 1 clic
    wb = load_workbook('Modèle de ce qui est attendu.xlsx')
    ws = wb['03']
    
    # Exemple : on remplit la ligne du mois
    ligne = ws.max_row + 1
    ws[f'B{ligne}'] = data['date']
    ws[f'C{ligne}'] = data['montant_ht']
    ws[f'D{ligne}'] = data['retenue']
    # … (je remplis tout le tableau en 15 lignes)
    
    # Génère le PDF
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"Certificat_{data['entreprise']}.xlsx")

app.run(host='0.0.0.0', port=8000)
